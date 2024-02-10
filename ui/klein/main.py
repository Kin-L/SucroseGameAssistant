# -*- coding:gbk -*-
import os
import webbrowser
from .list import KleinList
from .stack import KleinStack
from ui.element.control import *
from tools.environment import env
from tools.system import check_path


# 原神模组设置窗口
class Klein:
    def __init__(self, stack, icon, main):
        self.main = main
        self.widget_klein = Widget()
        stack.addWidget(self.widget_klein)
        self.button_klein = (
            PicButton(icon, (55, 0, 50, 50),
                      r"assets\kleins\picture\klein-icon.png", (50, 50)))
        self.list = None
        self.set = None

    def load_window(self):
        self.list = KleinList(self.widget_klein, (0, 0, 215, 515))
        self.set = KleinStack(self.widget_klein, (225, 0, 410, 515))
        self.list.set_klein.clicked.connect(lambda: self.set.stack.setCurrentIndex(0))
        self.list.set_fight.clicked.connect(lambda: self.set.stack.setCurrentIndex(1))
        self.list.set_disp.clicked.connect(lambda: self.set.stack.setCurrentIndex(2))
        self.list.set_review.clicked.connect(lambda: self.set.stack.setCurrentIndex(3))
        self.list.set_market.clicked.connect(lambda: self.set.stack.setCurrentIndex(4))
        self.list.set_recruit.clicked.connect(lambda: self.set.stack.setCurrentIndex(5))
        self.list.set_reward.clicked.connect(lambda: self.set.stack.setCurrentIndex(6))
        self.list.set_network.clicked.connect(lambda: self.set.stack.setCurrentIndex(7))
        self.list.set_mail.clicked.connect(lambda: self.set.stack.setCurrentIndex(8))
        self.list.set_roll.clicked.connect(lambda: self.set.stack.setCurrentIndex(9))
        self.set.button_arrange.clicked.connect(self.arrange_roll)
        self.set.button_open_roll.clicked.connect(self.open_roll)
        self.set.button_gift.clicked.connect(self.gift_inquiry)
        self.set.button_wiki.clicked.connect(self.open_wiki)
        self.set.button_history.clicked.connect(self.open_recruit_history)
        self.set.button_directory.clicked.connect(self.open_recruit_directory)
        Line(self.widget_klein, (215, 5, 3, 505), False)

    def load_run(self, run):
        _dir = {
            "server": 0,
            "game": "",
            "BGI": ""
        }
        _dir.update(run)
        self.set.combo_server.setCurrentIndex(_dir["server"])
        self.set.line_start.setText(_dir["game"])
        self.set.line_start.setSelection(0, 0)

    def get_run(self):
        _dir = {
            "server": self.set.combo_server.currentIndex(),
            "game": check_path(self.set.line_start.text())}
        return _dir

    def input_config(self, _dir):
        config = {
            "模块": 1,
            "静音": False,
            "关闭软件": False,
            "完成后": 0,
            "SGA关闭": False,
            "功能0": False,
            "功能1": False,
            "功能2": False,
            "功能3": False,
            "功能4": False,
            "功能5": False,
            "功能6": False,
            "功能7": False,
            "功能8": False,
            "再次重游": False,
            "作战关卡": 0,
            "再次采购": False,
            "采购0": [0, 0, 0],
            "采购1": [1, 0, 0],
            "采购2": [2, 0, 0],
            "采购3": [3, 0, 0],
            "采购4": [4, 0, 0],
            "采购5": [5, 0, 0],
            "回顾": 35.0,
            "使用加速": False,
            "访募金额": 0
            }
        config.update(_dir)
        self.set.independent.check_mute.setChecked(config["静音"])
        self.set.independent.check_kill_game.setChecked(config["关闭软件"])
        self.set.independent.combo_after.setCurrentIndex(config["完成后"])
        self.set.independent.check_kill_sga.setChecked(config["SGA关闭"])

        self.list.check_fight.setChecked(config["功能0"])
        self.list.check_disp.setChecked(config["功能1"])
        self.list.check_review.setChecked(config["功能2"])
        self.list.check_market.setChecked(config["功能3"])
        self.list.check_recruit.setChecked(config["功能4"])
        self.list.check_reward.setChecked(config["功能5"])
        self.list.check_network.setChecked(config["功能6"])
        self.list.check_mail.setChecked(config["功能7"])
        self.list.check_roll.setChecked(config["功能8"])

        self.set.re_fight.setChecked(config["再次重游"])
        self.set.mat.setCurrentIndex(config["作战关卡"])

        self.set.check_redisp.setChecked(config["再次采购"])
        self.set.mat0.setCurrentIndex(config["采购0"][0])
        self.set.mat1.setCurrentIndex(config["采购1"][0])
        self.set.mat2.setCurrentIndex(config["采购2"][0])
        self.set.mat3.setCurrentIndex(config["采购3"][0])
        self.set.mat4.setCurrentIndex(config["采购4"][0])
        self.set.mat5.setCurrentIndex(config["采购5"][0])

        self.set.fund0.setCurrentIndex(config["采购0"][1])
        self.set.fund1.setCurrentIndex(config["采购1"][1])
        self.set.fund2.setCurrentIndex(config["采购2"][1])
        self.set.fund3.setCurrentIndex(config["采购3"][1])
        self.set.fund4.setCurrentIndex(config["采购4"][1])
        self.set.fund5.setCurrentIndex(config["采购5"][1])

        self.set.plan0.setCurrentIndex(config["采购0"][2])
        self.set.plan1.setCurrentIndex(config["采购1"][2])
        self.set.plan2.setCurrentIndex(config["采购2"][2])
        self.set.plan3.setCurrentIndex(config["采购3"][2])
        self.set.plan4.setCurrentIndex(config["采购4"][2])
        self.set.plan5.setCurrentIndex(config["采购5"][2])

        self.set.num_box_review.setValue(config["回顾"])

        self.set.check_accelerate.setChecked(config["使用加速"])
        self.set.recruit_plan.setCurrentIndex(config["访募金额"])

    def output_config(self):
        config = dict()
        config["模块"] = 1
        config["静音"] = self.set.independent.check_mute.isChecked()
        config["关闭软件"] = self.set.independent.check_kill_game.isChecked()
        config["完成后"] = self.set.independent.combo_after.currentIndex()
        config["SGA关闭"] = self.set.independent.check_kill_sga.isChecked()

        config["功能0"] = self.list.check_fight.isChecked()
        config["功能1"] = self.list.check_disp.isChecked()
        config["功能2"] = self.list.check_review.isChecked()
        config["功能3"] = self.list.check_market.isChecked()
        config["功能4"] = self.list.check_recruit.isChecked()
        config["功能5"] = self.list.check_reward.isChecked()
        config["功能6"] = self.list.check_network.isChecked()
        config["功能7"] = self.list.check_mail.isChecked()
        config["功能8"] = self.list.check_roll.isChecked()

        config["再次重游"] = self.set.re_fight.isChecked()
        config["作战关卡"] = self.set.mat.currentIndex()

        config["再次采购"] = self.set.check_redisp.isChecked()
        config["采购0"] = [self.set.mat0.currentIndex(),
                         self.set.fund0.currentIndex(),
                         self.set.plan0.currentIndex()]
        config["采购1"] = [self.set.mat1.currentIndex(),
                         self.set.fund1.currentIndex(),
                         self.set.plan1.currentIndex()]
        config["采购2"] = [self.set.mat2.currentIndex(),
                         self.set.fund2.currentIndex(),
                         self.set.plan2.currentIndex()]
        config["采购3"] = [self.set.mat3.currentIndex(),
                         self.set.fund3.currentIndex(),
                         self.set.plan3.currentIndex()]
        config["采购4"] = [self.set.mat4.currentIndex(),
                         self.set.fund4.currentIndex(),
                         self.set.plan4.currentIndex()]
        config["采购5"] = [self.set.mat5.currentIndex(),
                         self.set.fund5.currentIndex(),
                         self.set.plan5.currentIndex()]

        config["回顾"] = self.set.num_box_review.value()
        config["使用加速"] = self.set.check_accelerate.isChecked()
        config["访募金额"] = self.set.recruit_plan.currentIndex()
        return config

    def open_roll(self):
        self.main.indicate("", 1)
        os.startfile(env.workdir + "/personal/kleins/roll")
        self.main.indicate("打开文件夹: 联络记录", 3)

    def open_recruit_directory(self):
        self.main.indicate("", 1)
        os.startfile(env.workdir + "/personal/kleins/recruit")
        self.main.indicate("打开文件: 访募历史", 3)

    def open_recruit_history(self):
        self.main.indicate("", 1)
        os.startfile(env.workdir + "/personal/kleins/recruit/history.txt")
        self.main.indicate("打开文件夹: 访募截图", 3)

    def gift_inquiry(self):
        self.main.indicate("", 1)
        webbrowser.open("https://www.bilibili.com/read/cv24639360/?from=search&spm_id_from=333.337.0.0")
        self.main.indicate("打开网页: 礼物认可度对照表", 3)

    def open_wiki(self):
        self.main.indicate("", 1)
        webbrowser.open("https://wiki.biligame.com/kelaiyinshe/%E8%88%8D%E5%8F%8B%E5%9B%BE%E9%89%B4")
        self.main.indicate("打开网页: 环行旅舍 BWIKI", 3)

    def arrange_roll(self):
        import json
        self.main.indicate("", 1)
        his_path = "personal/kleins/roll/history.json"
        if os.path.exists(his_path):
            with open(his_path, 'r', encoding='utf-8') as m:
                _dir = json.load(m)
        else:
            self.main.indicate("尚无已识别的联络记录", 3)
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        _excel = Workbook()
        sheet0 = _excel.create_sheet("总览")
        sheet0["A1"] = "抽卡总览"
        sheet0["A2"] = "定向联络"
        sheet0["B2"] = "N"
        sheet0["C2"] = "R"
        sheet0["D2"] = "SR"
        sheet0["E2"] = "SRR"
        sheet0["F2"] = "总抽数"
        sheet0["G2"] = "卡池水位"
        sheet0["A3"] = "计数"

        sheet0["A5"] = "常态联络"
        sheet0["B5"] = "N"
        sheet0["C5"] = "R"
        sheet0["D5"] = "SR"
        sheet0["E5"] = "SSR"
        sheet0["F5"] = "总抽数"
        sheet0["G5"] = "卡池水位"
        sheet0["A6"] = "计数"

        sheet0["A8"] = "初始联络"
        sheet0["B8"] = "N"
        sheet0["C8"] = "R"
        sheet0["D8"] = "SR"
        sheet0["E8"] = "SRR"
        sheet0["F8"] = "总抽数"
        sheet0["G8"] = "卡池水位"
        sheet0["A9"] = "计数"

        sheet0["A11"] = "限定联络"
        sheet0["B11"] = "N"
        sheet0["C11"] = "R"
        sheet0["D11"] = "SR"
        sheet0["E11"] = "SRR"
        sheet0["F11"] = "总抽数"
        sheet0["G11"] = "卡池水位"
        sheet0["A12"] = "计数"

        fon1 = Font(name='等线', size=14, bold=True)
        fon2 = Font(name='宋体', size=12)
        fon_n = Font(name='宋体', size=12, color="6C6C6C")
        fon_r = Font(name='宋体', size=12, color="3374F8")
        fon_sr = Font(name='宋体', size=12, color="7E30FF", bold=True)
        fon_srr = Font(name='宋体', size=12, color="FFC332", bold=True)
        al = Alignment(horizontal='center', vertical='center')
        sheet0["A1"].font = fon1

        sheet0["A2"].font = fon1
        sheet0["B2"].font = fon1
        sheet0["C2"].font = fon1
        sheet0["D2"].font = fon1
        sheet0["E2"].font = fon1
        sheet0["F2"].font = fon1
        sheet0["G2"].font = fon1
        sheet0["A3"].font = fon1
        sheet0["B3"].font = fon_n
        sheet0["C3"].font = fon_r
        sheet0["D3"].font = fon_sr
        sheet0["E3"].font = fon_srr
        sheet0["F3"].font = fon2
        sheet0["G3"].font = fon2

        sheet0["A5"].font = fon1
        sheet0["B5"].font = fon1
        sheet0["C5"].font = fon1
        sheet0["D5"].font = fon1
        sheet0["E5"].font = fon1
        sheet0["F5"].font = fon1
        sheet0["G5"].font = fon1
        sheet0["A6"].font = fon1
        sheet0["B6"].font = fon_n
        sheet0["C6"].font = fon_r
        sheet0["D6"].font = fon_sr
        sheet0["E6"].font = fon_srr
        sheet0["F6"].font = fon2
        sheet0["G6"].font = fon2

        sheet0["A8"].font = fon1
        sheet0["B8"].font = fon1
        sheet0["C8"].font = fon1
        sheet0["D8"].font = fon1
        sheet0["E8"].font = fon1
        sheet0["F8"].font = fon1
        sheet0["G8"].font = fon1
        sheet0["A9"].font = fon1
        sheet0["B9"].font = fon_n
        sheet0["C9"].font = fon_r
        sheet0["D9"].font = fon_sr
        sheet0["E9"].font = fon_srr
        sheet0["F9"].font = fon2
        sheet0["G9"].font = fon2
        
        sheet0["A11"].font = fon1
        sheet0["B11"].font = fon1
        sheet0["C11"].font = fon1
        sheet0["D11"].font = fon1
        sheet0["E11"].font = fon1
        sheet0["F11"].font = fon1
        sheet0["G11"].font = fon1
        sheet0["A12"].font = fon1
        sheet0["B12"].font = fon_n
        sheet0["C12"].font = fon_r
        sheet0["D12"].font = fon_sr
        sheet0["E12"].font = fon_srr
        sheet0["F12"].font = fon2
        sheet0["G12"].font = fon2
        sheet0.column_dimensions['A'].width = 15
        sheet0.column_dimensions['B'].width = 10
        sheet0.column_dimensions['C'].width = 10
        sheet0.column_dimensions['D'].width = 10
        sheet0.column_dimensions['E'].width = 10
        sheet0.column_dimensions['F'].width = 15
        sheet0.column_dimensions['G'].width = 15
        sheet0["A1"].alignment = al
        sheet0["A1"].alignment = al
        sheet0["A2"].alignment = al
        sheet0["B2"].alignment = al
        sheet0["C2"].alignment = al
        sheet0["D2"].alignment = al
        sheet0["E2"].alignment = al
        sheet0["F2"].alignment = al
        sheet0["G2"].alignment = al
        sheet0["A3"].alignment = al
        sheet0["B3"].alignment = al
        sheet0["C3"].alignment = al
        sheet0["D3"].alignment = al
        sheet0["E3"].alignment = al
        sheet0["F3"].alignment = al
        sheet0["G3"].alignment = al
        sheet0["A5"].alignment = al
        sheet0["B5"].alignment = al
        sheet0["C5"].alignment = al
        sheet0["D5"].alignment = al
        sheet0["E5"].alignment = al
        sheet0["F5"].alignment = al
        sheet0["G5"].alignment = al
        sheet0["A6"].alignment = al
        sheet0["B6"].alignment = al
        sheet0["C6"].alignment = al
        sheet0["D6"].alignment = al
        sheet0["E6"].alignment = al
        sheet0["F6"].alignment = al
        sheet0["G6"].alignment = al
        sheet0["A8"].alignment = al
        sheet0["B8"].alignment = al
        sheet0["C8"].alignment = al
        sheet0["D8"].alignment = al
        sheet0["E8"].alignment = al
        sheet0["F8"].alignment = al
        sheet0["G8"].alignment = al
        sheet0["A9"].alignment = al
        sheet0["B9"].alignment = al
        sheet0["C9"].alignment = al
        sheet0["D9"].alignment = al
        sheet0["E9"].alignment = al
        sheet0["F9"].alignment = al
        sheet0["G9"].alignment = al
        
        sheet0["A11"].alignment = al
        sheet0["B11"].alignment = al
        sheet0["C11"].alignment = al
        sheet0["D11"].alignment = al
        sheet0["E11"].alignment = al
        sheet0["F11"].alignment = al
        sheet0["G11"].alignment = al
        sheet0["A12"].alignment = al
        sheet0["B12"].alignment = al
        sheet0["C12"].alignment = al
        sheet0["D12"].alignment = al
        sheet0["E12"].alignment = al
        sheet0["F12"].alignment = al
        sheet0["G12"].alignment = al
        sheet0.row_dimensions[1].height = 18
        sheet0.row_dimensions[2].height = 18
        sheet0.row_dimensions[3].height = 18
        sheet0.row_dimensions[5].height = 18
        sheet0.row_dimensions[6].height = 18
        sheet0.row_dimensions[8].height = 18
        sheet0.row_dimensions[9].height = 18
        sheet0.row_dimensions[11].height = 18
        sheet0.row_dimensions[12].height = 18
        _count = []
        for i in ["定向联络", "常态联络", "初始联络", "限定联络"]:
            _sheet = _excel.create_sheet(i)
            _sheet["A1"] = "时间"
            _sheet["B1"] = "卡池"
            _sheet["C1"] = "稀有度"
            _sheet["D1"] = "角色名"
            _sheet["E1"] = "总次数"
            _sheet["F1"] = "保底水位"
            _sheet.column_dimensions['A'].width = 23
            _sheet.column_dimensions['B'].width = 20
            _sheet.column_dimensions['C'].width = 10
            _sheet.column_dimensions['D'].width = 15
            _sheet.column_dimensions['E'].width = 10
            _sheet.column_dimensions['F'].width = 10
            _sheet.row_dimensions[1].height = 20
            _sheet["A1"].font = fon1
            _sheet["B1"].font = fon1
            _sheet["C1"].font = fon1
            _sheet["D1"].font = fon1
            _sheet["E1"].font = fon1
            _sheet["F1"].font = fon1
            _sheet["A1"].alignment = al
            _sheet["B1"].alignment = al
            _sheet["C1"].alignment = al
            _sheet["D1"].alignment = al
            _sheet["E1"].alignment = al
            _sheet["F1"].alignment = al
            _list = _dir[i]
            n_a = 1
            n_sr = 0
            n_ssr = 0
            count = [0, 0, 0, 0, 0, 0, [], []]
            for [t, p, v, c] in _list:
                n_a += 1
                n_sr += 1
                n_ssr += 1
                print(t, p, v, c, n_a)
                _a = _sheet[f"A{n_a}"]
                _b = _sheet[f"B{n_a}"]
                _c = _sheet[f"C{n_a}"]
                _d = _sheet[f"D{n_a}"]
                _e = _sheet[f"E{n_a}"]
                _f = _sheet[f"F{n_a}"]
                _sheet[f"A{n_a}"] = t
                _sheet[f"B{n_a}"] = p
                _sheet[f"C{n_a}"] = v
                _sheet[f"D{n_a}"] = c
                _sheet[f"E{n_a}"] = n_a - 1
                _sheet[f"F{n_a}"] = n_ssr
                _sheet[f"A{n_a}"].alignment = al
                _sheet[f"B{n_a}"].alignment = al
                _sheet[f"C{n_a}"].alignment = al
                _sheet[f"D{n_a}"].alignment = al
                _sheet[f"E{n_a}"].alignment = al
                _sheet[f"F{n_a}"].alignment = al
                if v == "N":
                    _fon = fon_n
                    count[0] += 1
                elif v == "R":
                    _fon = fon_r
                    count[1] += 1
                elif v == "SR":
                    _fon = fon_sr
                    count[2] += 1
                    count[6] += [f"{c}({n_sr})"]
                    n_sr = 0
                elif v == "SSR":
                    _fon = fon_srr
                    count[3] += 1
                    count[7] += [f"{c}({n_ssr})"]
                    n_ssr = 0
                else:
                    print("稀有度异常")
                    continue
                _sheet[f"A{n_a}"].font = _fon
                _sheet[f"B{n_a}"].font = _fon
                _sheet[f"C{n_a}"].font = _fon
                _sheet[f"D{n_a}"].font = _fon
                _sheet[f"E{n_a}"].font = fon2
                _sheet[f"F{n_a}"].font = fon2
                _sheet.row_dimensions[n_a].height = 18
            count[4] = n_a-1
            count[5] = n_ssr
            _count += [count]
        sheet0["B3"] = _count[0][0]
        sheet0["C3"] = _count[0][1]
        sheet0["D3"] = _count[0][2]
        sheet0["E3"] = _count[0][3]
        sheet0["F3"] = _count[0][4]
        sheet0["G3"] = _count[0][5]

        _list = _count[0][6]
        _str = ""
        for i in _list:
            _str += i + " "
        sheet0["I2"] = _str
        _list = _count[0][7]
        _str = ""
        for i in _list:
            _str += i + " "
        sheet0["I3"] = _str

        sheet0["B6"] = _count[1][0]
        sheet0["C6"] = _count[1][1]
        sheet0["D6"] = _count[1][2]
        sheet0["E6"] = _count[1][3]
        sheet0["F6"] = _count[1][4]
        sheet0["G6"] = _count[1][5]
        _list = _count[1][6]
        _str = ""
        for i in _list:
            _str += i + " "
        sheet0["I5"] = _str

        _list = _count[1][7]
        _str = ""
        for i in _list:
            _str += i + " "
        sheet0["I6"] = _str

        sheet0["B9"] = _count[2][0]
        sheet0["C9"] = _count[2][1]
        sheet0["D9"] = _count[2][2]
        sheet0["E9"] = _count[2][3]
        sheet0["F9"] = _count[2][4]
        sheet0["G9"] = _count[2][5]
        _list = _count[2][6]
        _str = ""
        for i in _list:
            _str += i + " "
        sheet0["I8"] = _str

        _list = _count[2][7]
        _str = ""
        for i in _list:
            _str += i + " "
        sheet0["I9"] = _str

        sheet0["B12"] = _count[3][0]
        sheet0["C12"] = _count[3][1]
        sheet0["D12"] = _count[3][2]
        sheet0["E12"] = _count[3][3]
        sheet0["F12"] = _count[3][4]
        sheet0["G12"] = _count[3][5]
        _list = _count[3][6]
        _str = ""
        for i in _list:
            _str += i + " "
        sheet0["I11"] = _str

        _list = _count[3][7]
        _str = ""
        for i in _list:
            _str += i + " "
        sheet0["I12"] = _str
        _excel.remove(_excel['Sheet'])
        import time
        now = time.strftime("%Y-%m-%d %H-%M-%S", time.localtime())
        _path = f"personal/kleins/roll/环行旅舍联络记录 - {now}.xlsx"
        _excel.save(_path)
        self.main.indicate("联络记录已导出", 3)
